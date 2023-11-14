import os
import pandas as pd
import numpy as np
from PIL import Image
import csv
import re
import json
import paddlehub as hub
import cv2
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
from Minipic_Distinguish import kernel_minipic

ocr = hub.Module(name="chinese_ocr_db_crnn_server", enable_mkldnn=True)       # mkldnn加速仅在CPU下有效


class Main_core():
    # def __init__(self,dataframe,input_path,output_path):
    #     super(Main_core, self).__init__()
    #
    #     self.dataframe = dataframe
    #     self.input_path = input_path
    #     self.output_path = output_path

    def __init__(self):
        super(Main_core, self).__init__()
        #
        # self.dataframe = dataframe
        # self.input_path = input_path
        # self.output_path = output_path

    def parse_pic_to_excel_data(self , input_path):
        df = pd.DataFrame(columns=['序号', '代号', '名称', '数量', '材料', '单件', '总计', '备注'])
        self.input_path = input_path
        raw = cv2.imread(self.input_path, 1)
        gray = cv2.cvtColor(raw, cv2.COLOR_BGR2GRAY)
        binary = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, -5)
        rows, cols = binary.shape
        scale = 40
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (cols // scale, 1))
        eroded = cv2.erode(binary, kernel, iterations=1)
        dilated_col = cv2.dilate(eroded, kernel, iterations=1)
        scale = 20
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, rows // scale))
        eroded = cv2.erode(binary, kernel, iterations=1)
        dilated_row = cv2.dilate(eroded, kernel, iterations=1)
        bitwise_and = cv2.bitwise_and(dilated_col, dilated_row)

        # 标识表格
        merge = cv2.add(dilated_col, dilated_row)  # add是将横竖线叠加到一起

        ys, xs = np.where(bitwise_and > 0)  # np.where（condition）当条件满足时 ，只返回焦点的坐标 ，这里分别返回了36个交点的坐标

        # 纵坐标
        y_point_arr = []
        # 横坐标
        x_point_arr = []

        i = 0
        sort_x_point = np.sort(xs)
        for i in range(len(sort_x_point) - 1):
            if sort_x_point[i + 1] - sort_x_point[i] > 10:
                x_point_arr.append(sort_x_point[i])
            i = i + 1
        x_point_arr.append(sort_x_point[i])  # 要将最后一个点加入

        i = 0  # 计数器
        sort_y_point = np.sort(ys)
        # 这个步骤只将跳变的点加入到了数组当中
        for i in range(len(sort_y_point) - 1):
            if (sort_y_point[i + 1] - sort_y_point[i] > 10):
                y_point_arr.append(sort_y_point[i])
            i = i + 1
        y_point_arr.append(sort_y_point[i])  # 要将最后一个点加入

        difference = y_point_arr[1] - y_point_arr[0]
        new_y_point = y_point_arr[0] - difference
        y_point_arr.insert(0, new_y_point)  # 由于第一个点识别不准（涉及到卷积核），导致缺一行，所以手动添加上
        #####################识别图像中的横纵坐标点


        data = [[] for i in range(len(y_point_arr))]
        for i in range(len(y_point_arr) - 1):
            for j in range(len(x_point_arr) - 1):
                # 在分割时，第一个参数为y坐标，第二个参数为x坐标
                cell = raw[y_point_arr[i]:y_point_arr[i + 1], x_point_arr[j]:x_point_arr[j + 1]]
                if (j == 0 or j == 3):
                    km= kernel_minipic(cell)
                    text1 = km.core(cell)

                    # text1 = pytesseract.image_to_string(cell, lang="eng", config='--psm 6')
                    # text1 = re.findall(r'[^\*"/:?\\|<>″′‖() ,〈\n]', text1, re.S)
                    # text1 = "".join(text1)
                else:
                    result = ocr.recognize_text(images=[cell])
                    if (result[0]["data"] == []):
                        text1 = " "
                        text1 = str(text1)
                    else:
                        text1 = result[0]["data"][0]["text"]
                        text1 = str(text1)

                data[i].append(text1)
            df.loc[len(df.index)] = data[i]
        return df

    def make_excel(self, output_path , dataframe , values):
        '''

        :param output_path: 界面选择的文件输出路径
        :param dataframe: 上一个函数传入的表格内容
        :param values: 具体的文件输出名
        :return:
        '''
        self.dataframe = dataframe
        self.output_path = output_path
        self.values = values
        df_2 = self.dataframe

        for i in range(len(df_2)):
            if "5DQ" in df_2["代号"][i]:
                df_2["序号"][i] = 1
            elif "5TB" in df_2["代号"][i]:
                df_2["序号"][i] = 2
            elif "5BB" in df_2["代号"][i]:
                df_2["序号"][i] = 3
            elif "8DQ" in df_2["代号"][i]:
                df_2["序号"][i] = 4
            elif "8TB" in df_2["代号"][i]:
                df_2["序号"][i] = 5
            elif "GB" in df_2["代号"][i]:
                df_2["序号"][i] = 6
            else:
                df_2["序号"][i] = 7

        df_3 = df_2.sort_values(by="序号", ascending=True).reset_index().copy()
        # df_3.drop(['Unnamed: 0', "序号", "index"], axis=1)
        df_3.drop(["序号", "index"], axis=1)

        # excel 里面增加列
        # 代号列
        list_C = []
        for i in range(9, 33):
            list_C.append("C{}".format(i))
        for i in range(49, 73):
            list_C.append("C{}".format(i))
        # 名称列
        list_F = []
        for i in range(9, 33):
            list_F.append("F{}".format(i))
        for i in range(49, 73):
            list_F.append("F{}".format(i))
        # 数量列
        list_O = []
        for i in range(9, 33):
            list_O.append("O{}".format(i))
        for i in range(49, 73):
            list_O.append("O{}".format(i))
        # 备注列
        list_AD = []
        for i in range(9, 33):
            list_AD.append("AD{}".format(i))
        for i in range(49, 73):
            list_AD.append("AD{}".format(i))
        # “单件”列
        list_T = []
        for i in range(9, 33):
            list_T.append("T{}".format(i))
        for i in range(49, 73):
            list_T.append("T{}".format(i))
        # “材料-名称”列
        list_X = []
        for i in range(9, 33):
            list_X.append("X{}".format(i))
        for i in range(49, 73):
            list_X.append("X{}".format(i))
        # “总件”列
        list_V = []
        for i in range(9, 33):
            list_V.append("V{}".format(i))
        for i in range(49, 73):
            list_V.append("V{}".format(i))

        self.output_path = output_path
        file_name = '/{}.xlsx'.format(values) # 这块出现了问题
        output_path_combine = self.output_path + file_name

        shutil.copy('.\新编明细表.xlsx', output_path_combine)
        # shutil.copy（要复制的文件 ， 复制后的文件名和位置）
        wb = Workbook()
        file_name = output_path_combine

        # 读取Excel内容
        wb = load_workbook(filename=file_name)
        ### 这里花费了大量的时间
        sheet = wb['总装']
        # 修改Excel内容
        for i in range(len(df_3)):
            sheet[list_C[i]] = df_3["代号"][i]
            sheet[list_F[i]] = df_3["名称"][i]
            sheet[list_O[i]] = df_3["数量"][i]
            sheet[list_AD[i]] = df_3["备注"][i]
            sheet[list_T[i]] = df_3["单件"][i]
            sheet[list_X[i]] = df_3["材料"][i]
            sheet[list_V[i]] = df_3["总计"][i]

        wb.save(filename=file_name)

        return

